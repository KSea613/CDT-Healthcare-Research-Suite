import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def load_excel(filepath, sheet_name=0):
    """Load an Excel file into a pandas DataFrame using pandas/openpyxl engine."""
    df = pd.read_excel(filepath, sheet_name=sheet_name, engine="openpyxl")
    return df


def load_first_sheet_openpyxl(filepath):
    """Load the first worksheet from an Excel file using openpyxl and return a DataFrame.

    This function reads the first worksheet, treats the first non-empty row as headers,
    and returns a pandas DataFrame. Raises ValueError on invalid or unreadable files.
    """
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except InvalidFileException as e:
        raise ValueError("Invalid Excel file") from e
    except Exception as e:
        raise ValueError(f"Could not open file: {e}") from e

    sheets = wb.worksheets
    if not sheets:
        return pd.DataFrame()

    sheet = sheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()

    # Find first row with any non-None values to use as header
    header_row = None
    header_index = 0
    for idx, r in enumerate(rows):
        if any(cell is not None for cell in r):
            header_row = [str(c) if c is not None else "" for c in r]
            header_index = idx
            break

    if header_row is None:
        return pd.DataFrame()

    data_rows = rows[header_index + 1 :]
    df = pd.DataFrame(data_rows, columns=header_row)
    return df


def save_excel(df, filepath):
    """Save a pandas DataFrame to an Excel file."""
    df.to_excel(filepath, index=False, engine="openpyxl")
